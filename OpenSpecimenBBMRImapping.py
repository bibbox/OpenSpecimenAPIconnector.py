#! /bin/python3

from os_core.site import sites
from os_core.jsons import Json_factory
from os_core.users import users
from os_core.collecttion_protocol_registration import collection_protocol_registration
from os_core.collection_protocoll import collection_protocol

import json
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook


# URL to OpenSpecimen and Logindata
base_url = 'http://biobank-7-2.silicolab.bibbox.org/openspecimen/rest/ng'
auth = ('admin', 'Login@123')

#shhetnames:
bb_sheet = "eu_bbmri_eric_biobanks"
per_sheet = "eu_bbmri_eric_persons"
cp_sheet = "eu_bbmri_eric_collections"

# Load headers of BBMRI_ERIC Directory
template_file_name="empty_eric_duo.xlsx"
biobank_header = pd.read_excel(template_file_name, sheet_name = bb_sheet )
collection_header = pd.read_excel(template_file_name, sheet_name = cp_sheet)
person_header = pd.read_excel(template_file_name, sheet_name = per_sheet)
bbmri_file = pd.read_excel(template_file_name, sheet_name = None)



#Convert Headers to json-dict
collection_header_json=json.loads(collection_header.to_json())
biobank_header_json=json.loads(biobank_header.to_json())
person_header_json=json.loads(person_header.to_json())

#initialize Users,CollectionProtocols, Sites
protocols = collection_protocol(base_url = base_url, auth = auth)
user = users(base_url = base_url, auth = auth)
site = sites(base_url = base_url, auth = auth)

#Dict what now can't be done via OpenSpecimen extension Fields in persons
persons_extensions={
    "id":"bbmri-eric:conntactID:GR_TEST",
    "title_before_name":None,
    "title_after_name":None,
    "zip":"123456",
    "city":"City",
    "country":"GR",
    "collections":None
}

biobank_extensions ={
    "partner_charter_signed":'0',
    "head_title_before_name":None,
    "head_role":"PI",
    "contact_priority":"1"
}

# Mapping between OpenSpecimen keywords and bbmri keywords regarding persons
person_map = {
    "firstName":"first_name",
    "lastName":"last_name",
    "emailAddress":"email",
    "phoneNumber":"phone",
    "address":"address",
    "instituteName":"biobanks"
}

# Mapping between OpenSpecimen keyqords and BBMRI keywords regarding biobanks
biobank_map = {
    "name":"name"
}

#Mapping between Openspecimne keywords and BBMRI keywords regarding CollectionProtocols
collection_map= {
    "shortTitle":"acronym",
    "title":"name",

}
#load Openspecimen Collection Protocols
cp = protocols.get_collection_protocol(cpid = "18")

#extract User
OS_user_id = cp['principalInvestigator']['id']
OS_user = user.get_user(userId = OS_user_id)

#extractSite
OS_site_name = cp['cpSites'][0]['siteName']
OS_sites = site.get_all_sites()
for i in range(len(OS_sites)):
    if OS_sites[i]['name']==OS_site_name:
        OS_site_id=OS_sites[i]['id']
OS_site=site.get_site(siteid = OS_site_id)


#fill the bbmri_persons_dict
person_json=person_header_json
for key in persons_extensions:
    person_json[key]=persons_extensions[key]

for key in OS_user:
    if key in person_map.keys():
        person_json[person_map[key]]=OS_user[key]


#fill the BBBMRI Biobank dict
biobank_json = biobank_header_json
biobank_json['name'] = OS_site_name
attrs = OS_site['extensionDetail']['attrs']
for atr in attrs:
    id_ = atr["caption"].lower()
    id_ = id_.replace(" ", "_")
    if isinstance(atr['value'],list):
        for i in range(len(atr['value'])):
            string+=atr['value'][i]+', '
        string=string[0:-2]
    else:
        string=atr['value']
    biobank_json[id_]=string

person_json['id']=biobank_json['contact']

attrs=cp['extensionDetail']['attrs']
collection_json=collection_header_json


##fill json-dict, with OpenSpecimen standard fields
for key in cp:
    if key in collection_map.keys():
        collection_json[collection_map[key]]=cp[key]


for atr in attrs:
    id_ = atr["caption"].lower()
    id_ = id_.replace(" ", "_")
    if id_== "bbmri_collection_id":    
        id_="id"
    if isinstance(atr['value'],list):
        string = str('')
        for i in range(len(atr['value'])):
            string+=str(atr['value'][i])+ ', '
        string=string[0:-2]
    else:
        string=atr['value']
    collection_json[id_]=string
    collection_json[id_]=atr['value']

# Write the excel file
filename='empty_eric_duo (copy).xlsx'
df = pd.json_normalize(biobank_json)
with pd.ExcelWriter(filename, engine = 'openpyxl', mode = 'a') as writer:
    writer.book = load_workbook(filename)
    df.to_excel(writer, sheet_name = bb_sheet,  index = False)
df = pd.json_normalize(collection_json)
with pd.ExcelWriter(filename, engine = 'openpyxl', mode = 'a') as writer:
    writer.book = load_workbook(filename)
    df.to_excel(writer, sheet_name = cp_sheet,  index = False)
df = pd.json_normalize(person_json)
with pd.ExcelWriter(filename, engine = 'openpyxl', mode = 'a') as writer:
    writer.book = load_workbook(filename)
    df.to_excel(writer, sheet_name = per_sheet,  index = False)

#writer = pd.ExcelWriter('test.xlsx', engine = 'openpyxl', mode = 'a')

#df.to_excel(writer, sheet_name = bb_sheet, index =False, header = None)
#writer.save()

'''
#print(bbmri_file)
#df = pd.json_normalize(biobank_json)
#print(type(df))
#df.to_excel('test.xlsx', index=False,  sheet_name=bb_sheet)

bbmri_file[bb_sheet] = pd.json_normalize(biobank_json)
bbmri_file[per_sheet] = pd.json_normalize(person_json)
bbmri_file[cp_sheet] = pd.json_normalize(collection_json)

df = pd.json_normalize(bbmri_file)
print(type(df))
print(df.ndim)

#print(bbmri_file)

#df = pd.DataFrame.from_dict(bbmri_file)

#df.to_excel('test.xlsx', sheet_name = None, index = False)
'''
